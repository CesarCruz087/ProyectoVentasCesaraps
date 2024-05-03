from flask import Flask, render_template, request, jsonify, redirect, send_file
from BD import Estructura
import datetime
from dash import Dash, dcc, html
from dash.dependencies import Input, Output
import plotly.graph_objs as go
from BD.CRUD import Categoria, Detalles, Producto, Proveedor, Venta
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font
import os

app = Flask(__name__)
session = Estructura.get_session()

server = app
app_dash = Dash(__name__, server=server, url_base_pathname='/dashboard/')

app_dash.layout = html.Div(children=[
    html.H1(children='Dashboard de Ventas de Productos'),

    html.Div(children='''
        Selecciona un rango de fechas:
    '''),

    dcc.DatePickerRange(
        id='date-range-picker',
        start_date_placeholder_text="Fecha de inicio",
        end_date_placeholder_text="Fecha de fin",
        start_date='2024-04-01',
        end_date='2024-04-30'
    ),

    html.Div(id='producto-mas-vendido', style={'fontSize': 24, 'margin': '20px', 'padding': '10px', 'border': '2px solid black'}),
    
    html.Div(id='dinero-ventas', style={'fontSize': 24, 'margin': '20px', 'padding': '10px', 'border': '2px solid black'}),

    dcc.Graph(id='ventas-graph'),
])

@app_dash.callback(
    [Output('ventas-graph', 'figure'),
     Output('producto-mas-vendido', 'children'),
     Output('dinero-ventas', 'children')],
    [Input('date-range-picker', 'start_date'),
     Input('date-range-picker', 'end_date')]
)

def update_ventas(start_date, end_date):
    ventas_por_producto = session.query(Estructura.Producto.Nombre, func.sum(Estructura.Detalles.Cantidad)).join(Estructura.Detalles).join(Estructura.Venta).filter(
        Estructura.Venta.Fecha.between(start_date, end_date)
    ).group_by(Estructura.Producto.Nombre).all()

    productos = [venta[0] for venta in ventas_por_producto]
    cantidades = [venta[1] for venta in ventas_por_producto]

    fig = go.Figure(go.Bar(x=productos, y=cantidades))

    fig.update_layout(title='Cantidad de Ventas por Producto',
                      xaxis_title='Producto',
                      yaxis_title='Cantidad de Ventas',
                      template='plotly_dark')

    producto_mas_vendido = max(ventas_por_producto, key=lambda x: x[1])
    mensaje_producto_mas_vendido = f"El producto más vendido es: {producto_mas_vendido[0]} con {producto_mas_vendido[1]} unidades vendidas."
    total_ventas = session.query(func.sum(Estructura.Venta.Total)).filter(
        Estructura.Venta.Fecha.between(start_date, end_date)
    ).scalar()

    mensaje_dinero_ventas = f"El total de dinero de las ventas realizadas es: ${round(total_ventas, 2)}"

    return fig, html.Div(mensaje_producto_mas_vendido, style={'fontSize': 20}), html.Div(mensaje_dinero_ventas, style={'fontSize': 20})

@app.route('/')
def index():
    productos = Producto.list_by_existence()
    return render_template('principal.html', productos=productos)

@app.route('/get_producto/<int:id>')
def get_producto(id):
    prod = Producto.read_by_id(id)
    prod_json = {
        'Nombre': prod.Nombre,
        'Descripcion': prod.Descripcion,
        'Precio': prod.PrecioUnitario,
        'Cantidad': prod.CantidadDisponible
    }
    return jsonify(prod_json)

@app.route('/finalizar_compra', methods=['POST'])
def finalizar_compra():
    carrito = request.get_json().get('carrito')
    total = request.get_json().get('total')
    venta = Venta.create(datetime.datetime.now(), total, session)
    for fila in carrito:
        detalle = Detalles.create(venta.Id, fila.get('Id'), fila.get('Cantidad'), 0, session)
        prod = Producto.read_by_id(fila.get('Id'))
        Producto.update(prod, CantidadDisponible=prod.CantidadDisponible-fila.get('Cantidad'))
    return jsonify({'success': True})

@app.route('/productos')
def productos():
    cats = Categoria.list()
    provs = Proveedor.list()
    prods = Producto.list()
    return render_template('productos.html', cats = cats, provs = provs, prods = prods)

@app.route('/list_productos')
def list_productos():
    prods = Producto.list()
    prods_json = []
    for prod in prods:
        prods_json.append({
            'Id': prod.Id,
            'Nombre': prod.Nombre,
            'Descripcion': prod.Descripcion,
            'Precio': prod.PrecioUnitario,
            'Cantidad': prod.CantidadDisponible
        })
    return jsonify(prods_json)

@app.route('/productos_proveedor/<int:Id>')
def filtro_proveedor(Id):
    prov = Proveedor.read_by_id(Id)
    prods = Producto.list_by_proveedor(prov)
    prods_json = []
    for prod in prods:
        prods_json.append({
            'Id': prod.Id,
            'Nombre': prod.Nombre,
            'Descripcion': prod.Descripcion,
            'Precio': prod.PrecioUnitario,
            'Cantidad': prod.CantidadDisponible
        })
    return jsonify(prods_json)
    
@app.route('/productos_categoria/<int:Id>')
def filtro_categoria(Id):
    cat = Categoria.read_by_id(Id)
    prods = Producto.list_by_categoria(cat)
    prods_json = []
    for prod in prods:
        prods_json.append({
            'Id': prod.Id,
            'Nombre': prod.Nombre,
            'Descripcion': prod.Descripcion,
            'Precio': prod.PrecioUnitario,
            'Cantidad': prod.CantidadDisponible
        })
    return jsonify(prods_json)

@app.route('/actualizar_precio', methods=['POST'])
def actualizar_precio():
    id = request.get_json().get('id')
    precio = request.get_json().get('precio')
    prod = Producto.read_by_id(id)
    Producto.update(prod, PrecioUnitario=precio)
    return jsonify({'success': True})

@app.route('/actualizar_cantidad', methods=['POST'])
def actualizar_cantidad():
    id = request.get_json().get('id')
    cantidad = request.get_json().get('cantidad')
    prod = Producto.read_by_id(id)
    Producto.update(prod, CantidadDisponible=prod.CantidadDisponible+cantidad)
    return jsonify({'success': True})

@app.route('/agregar_producto')
def agregar_producto():
    provs = Proveedor.list()
    cats = Categoria.list()
    return render_template('crear_producto.html', provs = provs, cats = cats)

@app.route('/agregar_proveedor')
def agregar_proveedor():
    return render_template('crear_proveedor.html')

@app.route('/agregar_categoria')
def agregar_categoria():
    return render_template('crear_categoria.html')

@app.route('/guardar_producto', methods=['POST'])
def guardar_producto():
    nombre = request.form['nombre']
    descripcion = request.form['descripcion']
    precio_unitario = float(request.form['precio_unitario'])
    cantidad_disponible = int(request.form['cantidad_disponible'])
    categoria_id = int(request.form['categoria_id'])
    proveedor_id = int(request.form['proveedor_id'])

    Producto.create(nombre, descripcion, precio_unitario, cantidad_disponible, categoria_id, proveedor_id)

    return """
                <script>
                    alert("Se guardó correctamente")
                    window.location.href = "productos"
                </script>
            """

@app.route('/guardar_proveedor', methods=['POST'])
def guardar_proveedor():
    nombre = request.form['nombre']
    descripcion = request.form['descripcion']

    Proveedor.create(nombre, descripcion)

    return """
                <script>
                    alert("Se guardó correctamente")
                    window.location.href = "productos"
                </script>
            """

@app.route('/guardar_categoria', methods=['POST'])
def guardar_categoria():
    nombre = request.form['nombre']
    descripcion = request.form['descripcion']

    Categoria.create(nombre, descripcion)

    return """
                <script>
                    alert("Se guardó correctamente")
                    window.location.href = "productos"
                </script>
            """

@app.route("/dash")
def dash():
    return render_template('dash.html', dashboard_content=app_dash.index())

@app.route('/excel')
def excel():
    wb = Workbook()
    sheet = wb.active
    provs = Proveedor.list()
    sheet.title = 'Productos de proveedores'
    sheet['A1'] = 'Proveedor'
    sheet['B1'] = 'Productos'
    sheet['B2'] = 'Nombre'
    sheet['C2'] = 'Precio'
    sheet['D2'] = 'Cantidad'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['B2'].font = Font(bold=True)
    sheet['C2'].font = Font(bold=True)
    sheet['D2'].font = Font(bold=True)
    i=3
    for prov in provs:
        sheet['A'+str(i)] = prov.Nombre
        prods = Producto.list_by_proveedor(prov)
        i+=1
        for prod in prods:
            sheet['B'+str(i)] = prod.Nombre
            sheet['C'+str(i)] = prod.PrecioUnitario
            sheet['D'+str(i)] = prod.CantidadDisponible
            i+=1

    sheet = wb.create_sheet('Productos de categorias')
    cats = Categoria.list()
    sheet['A1'] = 'Categoria'
    sheet['B1'] = 'Productos'
    sheet['B2'] = 'Nombre'
    sheet['C2'] = 'Precio'
    sheet['D2'] = 'Cantidad'
    sheet['A1'].font = Font(bold=True)
    sheet['B1'].font = Font(bold=True)
    sheet['B2'].font = Font(bold=True)
    sheet['C2'].font = Font(bold=True)
    sheet['D2'].font = Font(bold=True)
    i=3
    for cat in cats:
        sheet['A'+str(i)] = cat.Nombre
        prods = Producto.list_by_categoria(cat)
        i+=1
        for prod in prods:
            sheet['B'+str(i)] = prod.Nombre
            sheet['C'+str(i)] = prod.PrecioUnitario
            sheet['D'+str(i)] = prod.CantidadDisponible
            i+=1

    sheet = wb.create_sheet('Ventas')
    ventas = Venta.list()
    sheet['A1'] = 'Venta'
    sheet['C1'] = 'Productos'
    sheet['A2'] = 'Fecha'
    sheet['B2'] = 'Total'
    sheet['C2'] = 'Nombre'
    sheet['D2'] = 'Cantidad comprada'
    sheet['E2'] = 'Costo'
    sheet['A1'].font = Font(bold=True)
    sheet['C1'].font = Font(bold=True)
    sheet['A2'].font = Font(bold=True)
    sheet['B2'].font = Font(bold=True)
    sheet['C2'].font = Font(bold=True)
    sheet['D2'].font = Font(bold=True)
    sheet['E2'].font = Font(bold=True)
    i=3
    for venta in ventas:
        sheet['A'+str(i)] = venta.Fecha
        sheet['B'+str(i)] = venta.Total
        detalles = Detalles.list_por_venta(venta)
        i+=1
        for detalle in detalles:
            prod = Producto.read_by_id(detalle.ProductoId)
            sheet['C'+str(i)] = prod.Nombre
            sheet['D'+str(i)] = detalle.Cantidad
            sheet['E'+str(i)] = detalle.Cantidad * prod.PrecioUnitario
            i+=1

    for col in sheet.columns:
        sheet.column_dimensions[col[0].column_letter].auto_size = True



    nombre_archivo = 'archivo_excel.xlsx'
    ruta_archivo = os.path.join(app.root_path, nombre_archivo)
    wb.save(ruta_archivo)

    return send_file(ruta_archivo, as_attachment=True)

        
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
